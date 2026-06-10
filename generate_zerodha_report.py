#!/usr/bin/env python3
"""
Zerodha Training Hub — Full Excel Report Generator
Fetches live data from Firebase + HTML, produces template-matched .xlsx with charts.
Usage: python3 ~/Desktop/generate_zerodha_report.py
"""

import json, re, urllib.request, os, sys
from datetime import datetime

# ── Dependencies ──────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference, Series
    from openpyxl.chart.label import DataLabelList
except ImportError:
    os.system("pip3 install openpyxl -q")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference, Series

# ── Config ─────────────────────────────────────────────────────────────────────
FIREBASE_URL = "https://zerodha-training-hub-default-rtdb.firebaseio.com"
HTML_PATH    = os.path.expanduser("~/Desktop/zerodha-training-hub.html")
OUT_PATH     = os.path.expanduser("~/Desktop/Zerodha_Training_Dashboard.xlsx")
REPORT_END   = "May 2026"
MONTHS_ORDER = ["December 2025","January 2026","February 2026","March 2026",
                "April 2026","May 2026"]
TRAINERS     = ["Gaurav Kumar","Raju Vernekar","Freeda Reshma Dsouza",
                "Harish Bhat","Girish A"]

# ── Colours (exact from template) ─────────────────────────────────────────────
TEAL         = "00BFA5"
DARK_TEAL    = "00A187"
DARK_BG      = "232840"
DARK_BG2     = "1B2033"
WHITE        = "FFFFFF"
GREEN_HDR    = "258229"
GREEN_VAL    = "43A047"
PURPLE_HDR   = "5D43E1"
PURPLE_VAL   = "7B61FF"
AMBER_HDR    = "D78805"
AMBER_VAL    = "F5A623"
BLUE_HDR     = "0047A2"
BLUE_VAL     = "1565C0"
LIGHT_TEXT   = "E8EBF0"
DIM_TEXT     = "546E7A"
DARK_TEXT    = "1F2937"
ALT_ROW      = "F5F7FA"

def fill(hex_rgb):
    return PatternFill("solid", fgColor=hex_rgb)

def font(hex_rgb, bold=False, size=10, name="Calibri"):
    return Font(color=hex_rgb, bold=bold, size=size, name=name)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def style(ws, ref, fill_=None, font_=None, align_=None, border_=None):
    c = ws[ref]
    if fill_:  c.fill   = fill_
    if font_:  c.font   = font_
    if align_: c.alignment = align_
    if border_:c.border = border_

def merge_style(ws, rng, fill_=None, font_=None, align_=None, border_=None):
    ws.merge_cells(rng)
    top_left = ws[rng.split(":")[0]]
    if fill_:  top_left.fill   = fill_
    if font_:  top_left.font   = font_
    if align_: top_left.alignment = align_
    if border_:top_left.border = border_

# ── Fetch data ─────────────────────────────────────────────────────────────────
def fetch_firebase(key):
    try:
        url = f"{FIREBASE_URL}/{key}.json"
        with urllib.request.urlopen(url, timeout=10) as r:
            data = json.loads(r.read())
        if isinstance(data, dict):
            return list(data.values())
        return data or []
    except Exception as e:
        print(f"  Firebase fetch '{key}' failed: {e}")
        return []

def parse_tr_activities():
    """Extract hardcoded TR.activities from the HTML file."""
    try:
        with open(HTML_PATH, "r", encoding="utf-8") as f:
            html = f.read()
        m = re.search(r'"activities"\s*:\s*(\[.*?\])\s*,\s*"kra"', html, re.DOTALL)
        if m:
            return json.loads(m.group(1))
    except Exception as e:
        print(f"  Could not parse HTML: {e}")
    return []

def parse_tr_monthly():
    """Extract TR.monthly_hours from the HTML file."""
    try:
        with open(HTML_PATH, "r", encoding="utf-8") as f:
            html = f.read()
        m = re.search(r'"monthly_hours"\s*:\s*(\{.*?\})\s*,\s*"activities"', html, re.DOTALL)
        if m:
            return json.loads(m.group(1))
    except: pass
    return {}

print("Fetching data…")
base_acts  = parse_tr_activities()
logs       = fetch_firebase("zh_logs")
if isinstance(logs, list):
    all_logs = [l for l in logs if isinstance(l, dict)]
else:
    all_logs = []
all_acts   = base_acts + all_logs
# Filter to report months only
report_set = set(MONTHS_ORDER)
all_acts   = [a for a in all_acts if a.get("month") in report_set]
all_acts.sort(key=lambda a: MONTHS_ORDER.index(a.get("month", MONTHS_ORDER[-1]))
              if a.get("month") in MONTHS_ORDER else 99)

tr_mhrs    = parse_tr_monthly()

# Build monthly hours (TR base + log additions)
def get_monthly_hours():
    import copy
    base = copy.deepcopy(tr_mhrs)
    for lg in all_logs:
        m = lg.get("month","")
        if m not in report_set: continue
        if m not in base: base[m] = {t:0 for t in TRAINERS}
        tr = lg.get("trainer","")
        hrs = float(lg.get("hours",0))
        for t in TRAINERS:
            if t.split()[0] in tr or t in tr:
                base[m][t] = base[m].get(t,0) + hrs
    return base

mhrs = get_monthly_hours()

# Collect all activity types
act_types = {}
for a in all_acts:
    t = a.get("type","Other")
    act_types[t] = act_types.get(t,0) + float(a.get("hours",0))

all_types_sorted = sorted(act_types.items(), key=lambda x: -x[1])

print(f"  {len(all_acts)} activities across {len(set(a['month'] for a in all_acts))} months")

# ── Create workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws_dash = wb.active
ws_dash.title = "Dashboard"

# Column widths (A–R)
col_widths = {
    "A":1.5,"B":12,"C":5,"D":12,"E":5,"F":12,"G":5,
    "H":12,"I":5,"J":12,"K":5,"L":8,"M":5,"N":8,"O":5,
    "P":5,"Q":5,"R":5
}
for col_letter, width in col_widths.items():
    ws_dash.column_dimensions[col_letter].width = width

# Row heights
for r, h in {1:12,2:36,3:18,4:14,5:8,6:32,7:12,
             8:20,9:52,10:20,11:10,12:22,13:52,14:20,
             15:10,16:10,17:26,18:22,25:26,26:18}.items():
    ws_dash.row_dimensions[r].height = h

# ── Row 1: Teal accent bar ──────────────────────────────────────────────────
for col in range(1, 19):
    c = ws_dash.cell(row=1, column=col)
    c.fill = fill(TEAL)

# ── Row 2: Main title ────────────────────────────────────────────────────────
ws_dash.merge_cells("A2:R2")
c = ws_dash["A2"]
c.value = "ZERODHA  ·  TRAINING & DEVELOPMENT TRACKER"
c.fill  = fill(TEAL)
c.font  = Font(color=WHITE, bold=True, size=22, name="Calibri")
c.alignment = align("left", "center")

# ── Row 3: Subtitle ──────────────────────────────────────────────────────────
ws_dash.merge_cells("A3:R3")
c = ws_dash["A3"]
c.value = f"Training Activity Dashboard   |   {MONTHS_ORDER[0]} – {REPORT_END}"
c.fill  = fill(DARK_BG)
c.font  = font(LIGHT_TEXT, size=10)
c.alignment = align("left", "center")

# ── Row 4: Team names ─────────────────────────────────────────────────────────
ws_dash.merge_cells("A4:R4")
c = ws_dash["A4"]
c.value = "Freeda Reshma Dsouza  ·  Gaurav Kumar  ·  Girish A  ·  Harish Bhat  ·  Raju Vernekar"
c.fill  = fill(DARK_BG)
c.font  = font(DIM_TEXT, size=9)
c.alignment = align("left", "center")

# ── Row 5: Spacer ────────────────────────────────────────────────────────────
for col in range(1, 19):
    ws_dash.cell(row=5, column=col).fill = fill(DARK_BG)

# ── Row 6: Filter ─────────────────────────────────────────────────────────────
ws_dash.merge_cells("B6:C6")
c = ws_dash["B6"]
c.value = "FILTER BY MONTH"
c.fill  = fill(DARK_BG)
c.font  = font(LIGHT_TEXT, bold=True, size=10)
c.alignment = align("center","center")

# D6 = the filter cell (single cell, changeable by user)
c = ws_dash["D6"]
c.value = "All Months"   # default; user types month name here
c.fill  = fill(TEAL)
c.font  = font(DARK_BG2, bold=True, size=12)
c.alignment = align("center","center")

ws_dash.merge_cells("E6:Q6")
c = ws_dash["E6"]
c.value = '← Type a month name (e.g. "May 2026") or "All Months" to filter KPI cards & charts'
c.fill  = fill(DARK_BG)
c.font  = font("F0C040", size=9)
c.alignment = align("left","center")

# ── Row 7: Spacer ────────────────────────────────────────────────────────────
for col in range(1, 19):
    ws_dash.cell(row=7, column=col).fill = fill(DARK_BG)

# ── KPI Cards (rows 8-10) ────────────────────────────────────────────────────
kpi_cols = [("B","C","00A187","00BFA5"),("D","E","258229","43A047"),
            ("F","G","5D43E1","7B61FF"),("H","I","D78805","F5A623")]
kpi_labels = ["TOTAL ACTIVITIES","TOTAL MAN HOURS","ONLINE HOURS","OFFLINE HOURS"]
kpi_subs   = ["sessions recorded","combined hours","virtual sessions","in-person sessions"]

# KPI formulas referencing CD sheet (which is filter-aware)
kpi_formulas = [
    "=CD!B42",   # Total Activities
    "=CD!B43",   # Total Man Hours
    "=CD!B44",   # Online Hours
    "=CD!B45",   # Offline Hours
]

for i, (c1, c2, bg_hdr, bg_val) in enumerate(kpi_cols):
    # Header
    ws_dash.merge_cells(f"{c1}8:{c2}8")
    cell = ws_dash[f"{c1}8"]
    cell.value = kpi_labels[i]
    cell.fill  = fill(bg_hdr)
    cell.font  = Font(color=WHITE, bold=True, size=8, name="Calibri")
    cell.alignment = align("center","center")
    # Value
    ws_dash.merge_cells(f"{c1}9:{c2}9")
    cell = ws_dash[f"{c1}9"]
    cell.value = kpi_formulas[i]
    cell.fill  = fill(bg_val)
    cell.font  = Font(color=WHITE, bold=True, size=26, name="Calibri")
    cell.alignment = align("center","center")
    # Subtitle
    ws_dash.merge_cells(f"{c1}10:{c2}10")
    cell = ws_dash[f"{c1}10"]
    cell.value = kpi_subs[i]
    cell.fill  = fill(bg_hdr)
    cell.font  = Font(color=WHITE, bold=False, size=8, name="Calibri")
    cell.alignment = align("center","center")

# ── Row 11: Spacer ───────────────────────────────────────────────────────────
for col in range(1, 19):
    ws_dash.cell(row=11, column=col).fill = fill(DARK_BG)

# ── Trainer Cards (rows 12-14) ───────────────────────────────────────────────
trainer_cols = [("B","C"),("D","E"),("F","G"),("H","I"),("J","K")]
trainer_short = ["GAURAV KUMAR","RAJU VERNEKAR","FREEDA RESHMA DSOUZA",
                 "HARISH BHAT","GIRISH A"]
# Trainer hour formulas (SUMIFS filtered by D6)
trainer_formulas = [f"=CD!B{47+i}" for i in range(5)]

for i, (c1, c2) in enumerate(trainer_cols):
    # Header
    ws_dash.merge_cells(f"{c1}12:{c2}12")
    cell = ws_dash[f"{c1}12"]
    cell.value = trainer_short[i]
    cell.fill  = fill(BLUE_HDR)
    cell.font  = Font(color=WHITE, bold=True, size=8, name="Calibri")
    cell.alignment = align("center","center")
    # Value
    ws_dash.merge_cells(f"{c1}13:{c2}13")
    cell = ws_dash[f"{c1}13"]
    cell.value = trainer_formulas[i]
    cell.fill  = fill(BLUE_VAL)
    cell.font  = Font(color=WHITE, bold=True, size=26, name="Calibri")
    cell.alignment = align("center","center")
    # Subtitle
    ws_dash.merge_cells(f"{c1}14:{c2}14")
    cell = ws_dash[f"{c1}14"]
    cell.value = "hours trained"
    cell.fill  = fill(BLUE_HDR)
    cell.font  = Font(color=WHITE, bold=False, size=8, name="Calibri")
    cell.alignment = align("center","center")

# ── Rows 15-16: Spacers ──────────────────────────────────────────────────────
for r in [15,16]:
    for col in range(1, 19):
        ws_dash.cell(row=r, column=col).fill = fill("F8FAFB")

# ── Row 17: Monthly Summary Header ──────────────────────────────────────────
ws_dash.merge_cells("B17:Q17")
c = ws_dash["B17"]
c.value = "MONTHLY HOURS SUMMARY"
c.fill  = fill(WHITE)
c.font  = Font(color=DARK_BG2, bold=True, size=11, name="Calibri")
c.alignment = align("left","center")

# ── Row 18: Monthly table column headers ─────────────────────────────────────
mth_hdr_cols = [("B","C","Month"),("D","E","Gaurav Kumar"),("F","G","Raju Vernekar"),
                ("H","I","Freeda Reshma D."),("J","K","Harish Bhat"),
                ("L","M","Girish A"),("N","O","Grand Total")]
for c1,c2,lbl in mth_hdr_cols:
    ws_dash.merge_cells(f"{c1}18:{c2}18") if c1!=c2 else None
    cell = ws_dash[f"{c1}18"]
    cell.value = lbl
    cell.fill  = fill(WHITE)
    cell.font  = Font(color=DIM_TEXT, bold=True, size=9, name="Calibri")
    cell.alignment = align("right" if lbl=="Grand Total" else "left","center")
    cell.border = Border(bottom=Side(style="thin", color="DDDDDD"))

# ── Monthly data rows (19+) ──────────────────────────────────────────────────
active_months = [m for m in MONTHS_ORDER if m in mhrs]
for idx, m in enumerate(active_months):
    row = 19 + idx
    ws_dash.row_dimensions[row].height = 20
    bg = WHITE if idx % 2 == 0 else ALT_ROW
    data = mhrs.get(m, {})
    grand = sum(data.get(t,0) for t in TRAINERS)

    # Month name
    ws_dash.merge_cells(f"B{row}:C{row}")
    cell = ws_dash[f"B{row}"]
    cell.value = m
    cell.fill  = fill(bg)
    cell.font  = Font(color=DARK_TEXT, bold=True, size=9, name="Calibri")
    cell.alignment = align("left","center")

    # Trainer values — use SUMIFS formulas referencing Data sheet
    tr_cols = ["D","F","H","J","L"]
    for j, (tc, trainer) in enumerate(zip(tr_cols, TRAINERS)):
        tc2 = chr(ord(tc)+1)
        ws_dash.merge_cells(f"{tc}{row}:{tc2}{row}")
        cell = ws_dash[f"{tc}{row}"]
        cell.value = f'=SUMIFS(Data!$I:$I,Data!$A:$A,"{m}",Data!$F:$F,"*{trainer.split()[0]}*")'
        cell.fill  = fill(bg)
        cell.font  = Font(color=DARK_TEXT, bold=False, size=9, name="Calibri")
        cell.alignment = align("right","center")
        cell.number_format = "0.00"

    # Grand Total
    ws_dash.merge_cells(f"N{row}:O{row}")
    cell = ws_dash[f"N{row}"]
    cell.value = f"=SUM(D{row},F{row},H{row},J{row},L{row})"
    cell.fill  = fill(bg)
    cell.font  = Font(color=PURPLE_VAL, bold=True, size=9, name="Calibri")
    cell.alignment = align("right","center")
    cell.number_format = "0.00"

# Grand total row
gt_row = 19 + len(active_months)
ws_dash.row_dimensions[gt_row].height = 22
ws_dash.merge_cells(f"B{gt_row}:C{gt_row}")
c = ws_dash[f"B{gt_row}"]
c.value = "GRAND TOTAL"
c.fill  = fill(DARK_BG)
c.font  = Font(color=TEAL, bold=True, size=9, name="Calibri")
c.alignment = align("left","center")

for j, tc in enumerate(["D","F","H","J","L"]):
    tc2 = chr(ord(tc)+1)
    ws_dash.merge_cells(f"{tc}{gt_row}:{tc2}{gt_row}")
    cell = ws_dash[f"{tc}{gt_row}"]
    cell.value = f"=SUM({tc}19:{tc}{gt_row-1})"
    cell.fill  = fill(DARK_BG)
    cell.font  = Font(color="F5A623", bold=True, size=9, name="Calibri")
    cell.alignment = align("right","center")
    cell.number_format = "0.00"

ws_dash.merge_cells(f"N{gt_row}:O{gt_row}")
c = ws_dash[f"N{gt_row}"]
c.value = f"=SUM(N19:N{gt_row-1})"
c.fill  = fill(DARK_BG)
c.font  = Font(color="F5A623", bold=True, size=9, name="Calibri")
c.alignment = align("right","center")
c.number_format = "0.00"

# ── Training Analytics section ────────────────────────────────────────────────
anl_row = gt_row + 2
ws_dash.row_dimensions[anl_row].height = 26
ws_dash.row_dimensions[anl_row+1].height = 18
ws_dash.merge_cells(f"B{anl_row}:Q{anl_row}")
c = ws_dash[f"B{anl_row}"]
c.value = "TRAINING ANALYTICS"
c.fill  = fill(WHITE)
c.font  = Font(color=DARK_BG2, bold=True, size=11, name="Calibri")
c.alignment = align("left","center")

chart_lbl_row = anl_row + 1
ws_dash.merge_cells(f"B{chart_lbl_row}:F{chart_lbl_row}")
ws_dash[f"B{chart_lbl_row}"].value = "Monthly Hours by Trainer"
ws_dash[f"B{chart_lbl_row}"].font  = Font(color=DIM_TEXT, size=9, name="Calibri")
ws_dash[f"B{chart_lbl_row}"].fill  = fill(WHITE)

ws_dash.merge_cells(f"G{chart_lbl_row}:J{chart_lbl_row}")
ws_dash[f"G{chart_lbl_row}"].value = "Online vs Offline"
ws_dash[f"G{chart_lbl_row}"].font  = Font(color=DIM_TEXT, size=9, name="Calibri")
ws_dash[f"G{chart_lbl_row}"].fill  = fill(WHITE)

ws_dash.merge_cells(f"K{chart_lbl_row}:R{chart_lbl_row}")
ws_dash[f"K{chart_lbl_row}"].value = "Hours by Activity Type"
ws_dash[f"K{chart_lbl_row}"].font  = Font(color=DIM_TEXT, size=9, name="Calibri")
ws_dash[f"K{chart_lbl_row}"].fill  = fill(WHITE)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — CD (Chart Data helper — filter-aware SUMIFS)
# ══════════════════════════════════════════════════════════════════════════════
ws_cd = wb.create_sheet("CD")

# Section 1: Monthly hours per trainer (rows 1-7)
ws_cd["A1"].value = "Month"
for j, t in enumerate(TRAINERS):
    ws_cd.cell(row=1, column=j+2).value = t

for i, m in enumerate(active_months):
    row = i + 2
    ws_cd.cell(row=row, column=1).value = m
    for j, t in enumerate(TRAINERS):
        first = t.split()[0]
        ws_cd.cell(row=row, column=j+2).value = (
            f'=SUMIFS(Data!$I:$I,Data!$A:$A,"{m}",Data!$F:$F,"*{first}*")'
        )

# Blank separator
# Section 2: Session type (rows 9-11)
ws_cd["A9"].value  = "Session Type"
ws_cd["B9"].value  = "Hours"
ws_cd["A10"].value = "Online"
ws_cd["A11"].value = "Offline"
ws_cd["B10"].value = '=IF(Dashboard!$D$6="All Months",SUMIF(Data!$G:$G,"Online",Data!$I:$I),SUMIFS(Data!$I:$I,Data!$A:$A,Dashboard!$D$6,Data!$G:$G,"Online"))'
ws_cd["B11"].value = '=IF(Dashboard!$D$6="All Months",SUMIF(Data!$G:$G,"Offline",Data!$I:$I),SUMIFS(Data!$I:$I,Data!$A:$A,Dashboard!$D$6,Data!$G:$G,"Offline"))'

# Section 3: Activity type breakdown (rows 14+)
ws_cd["A14"].value = "Activity"
ws_cd["B14"].value = "Hours"
for i, (act, _) in enumerate(all_types_sorted):
    ws_cd.cell(row=15+i, column=1).value = act
    ws_cd.cell(row=15+i, column=2).value = (
        f'=IF(Dashboard!$D$6="All Months",'
        f'SUMIF(Data!$B:$B,"{act}",Data!$I:$I),'
        f'SUMIFS(Data!$I:$I,Data!$A:$A,Dashboard!$D$6,Data!$B:$B,"{act}"))'
    )

act_end_row = 14 + len(all_types_sorted)

# Section 4: KPI summary for Dashboard cards (rows 42-51)
ws_cd["A42"].value = "Total Activities"
ws_cd["B42"].value = '=IF(Dashboard!$D$6="All Months",COUNTA(Data!$A:$A)-1,COUNTIFS(Data!$A:$A,Dashboard!$D$6))'
ws_cd["A43"].value = "Total Man Hours"
ws_cd["B43"].value = '=IF(Dashboard!$D$6="All Months",SUM(Data!$I:$I),SUMIFS(Data!$I:$I,Data!$A:$A,Dashboard!$D$6))'
ws_cd["A44"].value = "Online Hours"
ws_cd["B44"].value = "=B10"
ws_cd["A45"].value = "Offline Hours"
ws_cd["B45"].value = "=B11"

# Per-trainer totals for Dashboard cards (rows 47-51)
for i, t in enumerate(TRAINERS):
    first = t.split()[0]
    ws_cd.cell(row=47+i, column=1).value = t
    ws_cd.cell(row=47+i, column=2).value = (
        f'=IF(Dashboard!$D$6="All Months",'
        f'SUMIF(Data!$F:$F,"*{first}*",Data!$I:$I),'
        f'SUMIFS(Data!$I:$I,Data!$A:$A,Dashboard!$D$6,Data!$F:$F,"*{first}*"))'
    )

# ── CHARTS ────────────────────────────────────────────────────────────────────
chart_anchor_row = chart_lbl_row + 1  # first row for charts

# Chart 1: Stacked Bar — Monthly Hours by Trainer
bar1 = BarChart()
bar1.type    = "bar"
bar1.grouping = "stacked"
bar1.overlap  = 100
bar1.width   = 15
bar1.height  = 12
bar1.title   = None
bar1.legend.position = "b"
bar1.y_axis.title = "Hours"

trainer_colors = ["4472C4","ED7D31","A9D18E","FF0000","9966CC"]
cats = Reference(ws_cd, min_col=1, max_col=1, min_row=2, max_row=1+len(active_months))
for j, trainer in enumerate(TRAINERS):
    vals = Reference(ws_cd, min_col=j+2, max_col=j+2, min_row=2, max_row=1+len(active_months))
    s = Series(vals, title=trainer)
    s.graphicalProperties.solidFill = trainer_colors[j % len(trainer_colors)]
    bar1.series.append(s)
bar1.set_categories(cats)
ws_dash.add_chart(bar1, f"B{chart_anchor_row}")

# Chart 2: Pie — Online vs Offline
pie = PieChart()
pie.width  = 10
pie.height = 12
pie.title  = None
pie.legend.position = "b"
pie_vals  = Reference(ws_cd, min_col=2, min_row=10, max_row=11)
pie_cats  = Reference(ws_cd, min_col=1, min_row=10, max_row=11)
pie_s     = Series(pie_vals, title="Session Type")
pie.series.append(pie_s)
pie.dLbls = DataLabelList(showPercent=True, showVal=False, showCatName=False)
pie.set_categories(pie_cats)
ws_dash.add_chart(pie, f"G{chart_anchor_row}")

# Chart 3: Bar — Hours by Activity Type
n_types = len(all_types_sorted)
bar2 = BarChart()
bar2.type     = "bar"
bar2.grouping = "clustered"
bar2.width    = 17
bar2.height   = 12
bar2.title    = None
bar2.legend   = None
bar2.y_axis.title = "Hours"
bar2_vals = Reference(ws_cd, min_col=2, min_row=15, max_row=14+n_types)
bar2_cats = Reference(ws_cd, min_col=1, min_row=15, max_row=14+n_types)
bar2_s    = Series(bar2_vals, title="Hours")
bar2_s.graphicalProperties.solidFill = "00BFA5"
bar2.series.append(bar2_s)
bar2.set_categories(bar2_cats)
ws_dash.add_chart(bar2, f"K{chart_anchor_row}")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — DATA
# ══════════════════════════════════════════════════════════════════════════════
ws_data = wb.create_sheet("Data")
data_headers = ["Month","Type of Training","Training Details","Start Date",
                "Date of Completion","Trainer Name","Session Type",
                "Topic Covered","Total Man Hours"]
data_col_widths = [16,22,30,12,12,28,12,32,14]

for j, (hdr, w) in enumerate(zip(data_headers, data_col_widths)):
    col_letter = get_column_letter(j+1)
    ws_data.column_dimensions[col_letter].width = w
    cell = ws_data.cell(row=1, column=j+1, value=hdr)
    cell.fill  = fill(DARK_BG)
    cell.font  = Font(color=WHITE, bold=True, size=10, name="Calibri")
    cell.alignment = align("center","center")
    cell.border = thin_border()

for i, act in enumerate(all_acts):
    row = i + 2
    is_online = act.get("session","") == "Online"
    bg_color  = "E3F2FD" if is_online else "E8F5E9"
    vals = [
        act.get("month",""),
        act.get("type",""),
        act.get("details") or act.get("training_details",""),
        act.get("start") or act.get("date",""),
        act.get("end",""),
        act.get("trainer",""),
        act.get("session",""),
        act.get("topic",""),
        float(act.get("hours",0)),
    ]
    for j, v in enumerate(vals):
        cell = ws_data.cell(row=row, column=j+1, value=v)
        cell.fill  = fill(bg_color)
        cell.font  = Font(color="111827", size=10, name="Calibri")
        cell.alignment = align("left" if j < 8 else "right","center")
        cell.border = thin_border("EEEEEE")

# ══════════════════════════════════════════════════════════════════════════════
# SHEETS 4+  — PER-MONTH TRAINING LOGS
# ══════════════════════════════════════════════════════════════════════════════
def build_month_sheet(wb, month_name):
    acts = [a for a in all_acts if a.get("month") == month_name]
    short = month_name.split()[0][:3] + (" 25" if "2025" in month_name else " 26")
    ws = wb.create_sheet(short)

    col_widths = {"A":4,"B":22,"C":30,"D":12,"E":12,"F":28,"G":12,"H":32,"I":12}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Row 1 + 2: Title
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 10
    ws.row_dimensions[6].height = 22

    ws.merge_cells("A1:I1")
    ws["A1"].fill = fill(TEAL)

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = f"{month_name.upper()}  —  TRAINING LOG"
    c.fill  = fill(TEAL)
    c.font  = Font(color=WHITE, bold=True, size=16, name="Calibri")
    c.alignment = align("left","center")

    ws.merge_cells("A3:I3")
    c = ws["A3"]
    c.value = f"Zerodha Training & Development  ·  All activities for {month_name}"
    c.fill  = fill(DARK_BG)
    c.font  = font(LIGHT_TEXT, size=9)
    c.alignment = align("left","center")

    ws.merge_cells("A4:I4")
    c = ws["A4"]
    c.value = "← Use the sheet tabs at the bottom to navigate to a different month"
    c.fill  = fill(DARK_BG)
    c.font  = font(DIM_TEXT, size=9)
    c.alignment = align("left","center")

    for col in range(1,10):
        ws.cell(row=5, column=col).fill = fill("F8FAFB")

    # Row 6: Headers
    hdrs = ["#","Type of Training","Training Details","Start Date","Completed",
            "Trainer Name","Session","Topic Covered","Man Hours"]
    for j, h in enumerate(hdrs):
        cell = ws.cell(row=6, column=j+1, value=h)
        cell.fill  = fill(DARK_BG)
        cell.font  = Font(color=WHITE, bold=True, size=10, name="Calibri")
        cell.alignment = align("center" if j == 0 else "left","center")
        cell.border = thin_border()

    # Data rows
    for i, act in enumerate(acts):
        row = 7 + i
        ws.row_dimensions[row].height = 18
        is_online = act.get("session","") == "Online"
        bg = "E3F2FD" if is_online else "E8F5E9"
        row_vals = [
            i+1,
            act.get("type",""),
            act.get("details") or act.get("training_details",""),
            act.get("start") or act.get("date",""),
            act.get("end",""),
            act.get("trainer",""),
            act.get("session",""),
            act.get("topic",""),
            float(act.get("hours",0)),
        ]
        for j, v in enumerate(row_vals):
            cell = ws.cell(row=row, column=j+1, value=v)
            cell.fill  = fill(bg)
            cell.font  = Font(color="111827", size=10, name="Calibri")
            cell.alignment = align("center" if j==0 else ("right" if j==8 else "left"),"center")
            cell.border = thin_border("DDDDDD")

    # Total row
    total_row = 7 + len(acts)
    ws.row_dimensions[total_row].height = 22
    ws.merge_cells(f"A{total_row}:H{total_row}")
    c = ws[f"A{total_row}"]
    c.value = "TOTAL MAN HOURS THIS MONTH"
    c.fill  = fill(DARK_BG)
    c.font  = Font(color=TEAL, bold=True, size=10, name="Calibri")
    c.alignment = align("left","center")

    c = ws[f"I{total_row}"]
    c.value = f"=SUM(I7:I{total_row-1})" if acts else 0
    c.fill  = fill(DARK_BG)
    c.font  = Font(color="F5A623", bold=True, size=10, name="Calibri")
    c.alignment = align("right","center")
    c.number_format = "0.00"

    return ws

for m in active_months:
    build_month_sheet(wb, m)

# ── Save ────────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"\n✅  Saved: {OUT_PATH}")
print(f"   Sheets: Dashboard, CD, Data, {', '.join([m.split()[0][:3]+(' 25' if '2025' in m else ' 26') for m in active_months])}")
print(f"\nTo use the filter: open the file → Dashboard tab → change cell D6 to a month name")
print("(e.g. 'May 2026') — all KPI cards and charts update automatically.")
